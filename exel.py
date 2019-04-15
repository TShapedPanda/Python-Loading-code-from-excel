# First import Openpyxl and matplotlib
from openpyxl import load_workbook
from matplotlib import pyplot as plt
#Our data called hours contains 50 rows of age and hours 
workbook = load_workbook('hours.xlsx')
#mylistx placeholder for x values to be plotted which are ages
mylistx=[]
#mylisty placeholder for y vales to be plotted which are hours watching tv
mylisty=[]
#sheet grabs the active worksheet  , rows and columns
sheet = workbook.active
xtitle=sheet.cell(row=1, column=1).value
ytitle=sheet.cell(row=1, column=2).value
row_count = 52

#iterate through the data adding values to the list.
for i in range(2,row_count):
       mylistx2=[sheet.cell(row=i, column=1).value]
       mylisty2=[sheet.cell(row=i, column=2).value]
       mylistx=mylistx+mylistx2
       mylisty=mylisty+mylisty2
#Using matplotlib we are able to plot a basci line graph from the iterated list.       
plt.figure(1)   
plt.xlabel(xtitle)
plt.ylabel(ytitle)
plt.plot(mylistx,mylisty)
plt.title('TV hours by Age')
plt.show()
#Using matplotlib we are are able to plot a bar chart.
plt.figure(2)
plt.ylabel("Daily TV Hours watched")
plt.xlabel(xtitle)
plt.ylabel(ytitle)

plt.bar(mylistx,mylisty)
plt.show()

#just a comment
#just another comment
Print (mylistx2)
print(mylisty2)
